import os
import sys
import time
import openpyxl
import webbrowser
from google.genai import types
from pydantic import BaseModel, Field

# 載入主程式的設定與 API 用戶端
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from news_collector import EXCEL_PATH, client, generate_html_dashboard

# Pydantic 結構化輸出模型
class BackfillAnalysis(BaseModel):
    impact_analysis: str = Field(description="以『【AI 觀點】...』開頭，用一句話評估對筆電市場/華碩的關鍵影響與機會/風險，字數限制在 30-50 字內，極度精簡")

def analyze_impact_from_excel(topic, content):
    """
    呼叫 Gemini 2.5-flash，根據現有的標題與內容，生成華碩與筆電市場的影響分析
    """
    prompt = f"""
您是一位專業的科技與財經分析師。
目前我們有一篇新聞的標題為：「{topic}」。
其新聞內容/摘要為：「{content}」。

請站在專業產業分析師的角度，具體評估該事件對「整個筆記型電腦 (PC/NB) 市場」與「華碩 (ASUS)」的潛在影響、風險、機會或出貨動態。
要求：
- 必須以「【AI 觀點】」開頭。
- 只能用「一句話」評估，字數限制約 30-50 字，直陳核心點，極度精簡。

請嚴格遵守回傳的 JSON 格式 Schema，以繁體中文 (Traditional Chinese) 回答。
"""
    try:
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt,
            config={
                'response_mime_type': 'application/json',
                'response_schema': BackfillAnalysis,
            }
        )
        return response.parsed.impact_analysis
    except Exception as e:
        print(f"  [警訊] Gemini 影響分析生成失敗: {e}")
        return None

def main():
    print("=" * 60)
    print(" 啟動缺失 I 欄 (影響分析) 補齊工具")
    print(f" Excel 目標：{EXCEL_PATH}")
    print("=" * 60)
    
    if not os.path.exists(EXCEL_PATH):
        print(f"[錯誤] 找不到 Excel 檔案：{EXCEL_PATH}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "News" not in wb.sheetnames:
        print("[錯誤] Excel 中找不到 'News' 工作表。")
        sys.exit(1)
        
    ws = wb["News"]
    
    # 確保第 9 欄 (I 欄) 命名正確
    if ws.cell(row=1, column=9).value != '對於筆電市場/華碩的影響':
        print("  [資訊] 將第 9 欄 (I 欄) 命名為「對於筆電市場/華碩的影響」...")
        ws.cell(row=1, column=9, value='對於筆電市場/華碩的影響')

    max_row = ws.max_row
    print(f"目前 Excel 中共有 {max_row} 行資料。")
    
    # 掃描前 50 行
    check_limit = min(50, max_row)
    print(f"我們將掃描前 {check_limit} 行，為手動新增或缺少影響分析的新聞補齊...")
    
    backfilled_count = 0
    
    # 從第二行 (Header 下方) 開始掃描
    for row in range(2, check_limit + 1):
        topic_val = ws.cell(row=row, column=4).value # Column 4 是 Topic
        content_val = ws.cell(row=row, column=5).value # Column 5 是 Content
        impact_val = ws.cell(row=row, column=9).value # Column 9 是 I 欄 (對於筆電市場/華碩的影響)
        
        # 如果標題不為空，且 I 欄為空
        if topic_val and (not impact_val or str(impact_val).strip() in ["", "NaN", "nan"]):
            print(f"\n[掃描到空值] 發現缺少分析的新聞 (第 {row} 行)：")
            print(f"  標題: {topic_val[:40]}...")
            
            # 若無摘要內容，則以標題作為參考
            ref_content = content_val if content_val else topic_val
            
            print("  正在呼叫 Gemini 生成影響分析...")
            impact_analysis = analyze_impact_from_excel(topic_val, ref_content)
            
            if impact_analysis:
                ws.cell(row=row, column=9, value=impact_analysis)
                print(f"  [成功] 已填補影響分析：{impact_analysis}")
                backfilled_count += 1
                time.sleep(1) # 友善間隔，避免 429 限流
            else:
                print("  [警訊] 無法為此行生成影響分析。")

    if backfilled_count > 0:
        try:
            wb.save(EXCEL_PATH)
            print(f"\n[成功] 成功補齊了 {backfilled_count} 筆新聞的 I 欄資料！")
            
            # 重新生成 HTML
            local_html_path = r"D:\ASUS\Anti-NotebookLM\NEWS\index.html"
            generate_html_dashboard(EXCEL_PATH, local_html_path)
            
            # 自動開啟瀏覽器
            webbrowser.open(local_html_path)
            print("[成功] 已重新渲染網頁並自動在瀏覽器中開啟！")
        except PermissionError:
            print("\n" + "!"*60)
            print("[錯誤] 無法儲存 Excel 檔案！請先關閉您的 Excel 軟體，再重新執行此腳本。")
            print("!"*60 + "\n")
    else:
        # 如果沒有新增補齊，但使用者依然想刷新 HTML 的週數和過濾空白行，我們也可以主動幫他重新生成
        print("\n[資訊] 檢查完畢，前 50 行中沒有發現需要補齊影響分析的新聞。")
        print("重新渲染最新的 HTML 看板中...")
        local_html_path = r"D:\ASUS\Anti-NotebookLM\NEWS\index.html"
        generate_html_dashboard(EXCEL_PATH, local_html_path)
        webbrowser.open(local_html_path)
        print("[成功] 已更新 HTML 看板網頁並開啟！")

if __name__ == "__main__":
    main()
