import os
import sys
import time
import html

# 強制將標準輸出設為 utf-8，避免 Windows CMD 下的 cp950 編碼錯誤 (如簡體字或表情符號)
if sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        pass
import urllib.parse
import webbrowser
from datetime import datetime
import xml.etree.ElementTree as ET

# 檢查依賴套件是否安裝，若無則友善提示
missing_packages = []
try:
    import pandas as pd
except ImportError:
    missing_packages.append("pandas")
try:
    import openpyxl
except ImportError:
    missing_packages.append("openpyxl")
try:
    import requests
except ImportError:
    missing_packages.append("requests")
try:
    from bs4 import BeautifulSoup
except ImportError:
    missing_packages.append("beautifulsoup4")
try:
    from google import genai
    from google.genai import types
except ImportError:
    missing_packages.append("google-genai")
try:
    from pydantic import BaseModel, Field
except ImportError:
    missing_packages.append("pydantic")
try:
    from googlenewsdecoder import gnewsdecoder
except ImportError:
    missing_packages.append("googlenewsdecoder")
try:
    from dotenv import load_dotenv
except ImportError:
    missing_packages.append("python-dotenv")

if missing_packages:
    print("=" * 60)
    print("[錯誤] 偵測到尚未安裝執行此工具所需的 Python 函式庫。")
    print("請開啟命令提示字元 (CMD) 並執行以下指令進行安裝：")
    print(f"pip install {' '.join(missing_packages)}")
    print("=" * 60)
    sys.exit(1)

# 載入環境變數
load_dotenv()
api_key = os.environ.get("GEMINI_API_KEY")

if not api_key:
    print("=" * 60)
    print("[錯誤] 找不到 GEMINI_API_KEY 環境變數或 .env 檔案設定。")
    print("請至以下網址申請免費的 API Key:")
    print("https://aistudio.google.com/")
    print("並在工作區建立 .env 檔案填入：GEMINI_API_KEY=您的金鑰")
    print("=" * 60)
    sys.exit(1)

# 讀取 DIGITIMES 會員設定
DIGITIMES_USER = os.environ.get("DIGITIMES_USER")
DIGITIMES_PASSWORD = os.environ.get("DIGITIMES_PASSWORD")

# 初始化 Google GenAI 用戶端，使用最新的穩定版 gemini-2.5-flash
client = genai.Client(api_key=api_key)

# Excel 檔案路徑與設定
EXCEL_PATH = r"D:\ASUS\News & Report\News\News.xlsx"

# Google Alerts RSS 訂閱源設定 (方案一)
GOOGLE_ALERT_RSS_ECON = os.environ.get("GOOGLE_ALERT_RSS_ECON")


# 四大主題分類與對應 Google News 搜尋關鍵字 (加入排除投資、娛樂等字詞)
CATEGORIES = {
    "國際經濟": '("全球" OR "美國" OR "中國" OR "台灣" OR "日本" OR "英國" OR "法國" OR "德國" OR "義大利" OR "加拿大" OR "澳洲") 經濟 產業 OR 產品 -股票 -股價 -投資 -理財 -娛樂 -八卦 -影劇 -site:gov.tw -site:edu.tw -site:gov -site:edu',
    "PC產業": '("PC" OR "AIPC" OR "AI PC" OR "個人電腦" OR "筆記型電腦" OR "筆電") 市場 OR 產業 OR 產品 -股票 -股價 -投資 -理財 -娛樂 -八卦 -影劇',
    "上游廠商動態": '(Intel OR AMD OR Nvidia OR "記憶體" OR Micron OR 美光 OR SK海力士 OR 三星半導體 OR 晶圓代工 OR 台積電) 產業 OR 產品 -股票 -股價 -投資 -理財 -娛樂 -八卦 -影劇',
    "各品牌廠動態": '(Dell OR HP OR Lenovo OR Apple OR Acer OR MSI OR Samsung OR ASUS OR 華碩 OR 聯想 OR 戴爾 OR 惠普 OR 宏碁 OR 三星) PC OR 筆電 產業 OR 產品 -股票 -股價 -投資 -理財 -娛樂 -八卦 -影劇'
}

# 每個類別預設抓取篇數
MAX_ARTICLES_PER_CATEGORY = 3

# Pydantic 結構化輸出模型 (具備全文時使用)
class NewsItemAnalysis(BaseModel):
    title: str = Field(description="根據新聞內容提煉極簡短的主題標題(Topic)，讓讀者能馬上抓到重點，不要用原標題，限 5-15 字。")
    content_clean: str = Field(description="從網頁純文字中提取出的純淨新聞全文內文。應排除廣告、導覽列、側邊欄、版權宣告等雜訊")
    publish_time: str = Field(description="新聞的發布時間。如果是 HTML 中有提供日期時間請提取 it，格式如 YYYY-MM-DD HH:MM 或 YYYY/MM/DD")
    summary: str = Field(description="針對此篇新聞內容生成精簡的內文敘述，請精簡成 2-3 句話。")
    impact_analysis: str = Field(description="以『[AI觀點]』開頭，評估對筆電市場/華碩的關鍵影響，請縮短為 1-2 句話。")

# Pydantic 結構化輸出模型 (無內文、僅標題分析時使用)
class TitleOnlyAnalysis(BaseModel):
    title: str = Field(description="根據原新聞標題提煉極簡短的主題標題(Topic)，讓讀者能馬上抓到重點，不要用原標題，限 5-15 字。")
    summary: str = Field(description="根據標題生成背景說明，請精簡成 2-3 句話。")
    impact_analysis: str = Field(description="以『[AI觀點]』開頭，評估對筆電市場/華碩的關鍵影響，請縮短為 1-2 句話。")

def get_digitimes_session(user, password):
    """
    使用 requests.Session() 模擬登入 DIGITIMES 台灣版網站以維持會員 Cookie
    """
    if not user or not password:
        return None
        
    print(f"  [資訊] 偵測到 DIGITIMES 會員設定，嘗試為帳號 {user[:4]}... 進行模擬登入...")
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7"
    })
    
    login_url = "https://www.digitimes.com.tw/member/login.asp"
    try:
        # 先 GET 登入頁取得 cookie
        session.get(login_url, timeout=5)
        
        # 準備 POST payload
        payload = {
            "member_id": user,
            "member_pwd": password,
            "act": "login"
        }
        
        # 發送登入 POST
        response = session.post(login_url, data=payload, timeout=5)
        
        # 簡單驗證登入結果：如果內容中含有常見的登入失敗提示
        if "帳號不存在" in response.text or "密碼不正確" in response.text or "請輸入" in response.text:
            print("  [警訊] DIGITIMES 登入失敗：請確認 .env 中的帳密是否正確。將使用公開模式抓取。")
            return None
            
        print("  [成功] DIGITIMES 會員登入成功！已建立會員 Session。")
        return session
    except Exception as e:
        print(f"  [警訊] DIGITIMES 登入時發生異常，將使用公開模式抓取。錯誤: {e}")
        return None

def fetch_google_news_rss(query, max_results=3):
    """
    抓取 Google News RSS 並返回新聞的標題、臨時連結與發布時間
    """
    encoded_query = urllib.parse.quote(query)
    url = f"https://news.google.com/rss/search?q={encoded_query}&hl=zh-TW&gl=TW&ceid=TW:zh-Hant"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=5)
        response.raise_for_status()
        
        root = ET.fromstring(response.content)
        items = []
        for item in root.findall(".//item")[:max_results]:
            title = item.find("title").text
            link = item.find("link").text
            pub_date = item.find("pubDate").text
            items.append({
                "title": title,
                "link": link,
                "pub_date": pub_date
            })
        return items
    except Exception as e:
        print(f"  [警訊] 抓取 RSS 失敗 ({query}): {e}")
        return []

def fetch_google_alert_rss(rss_url, max_results=8):
    """
    抓取 Google 快訊 (Google Alerts) 的 RSS Feed，並返回新聞的標題、真實連結與發布時間。
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    try:
        response = requests.get(rss_url, headers=headers, timeout=10)
        response.raise_for_status()
        
        # Google 快訊 RSS 是 Atom 格式
        root = ET.fromstring(response.content)
        
        # 定義 Atom namespace
        ns = {'atom': 'http://www.w3.org/2005/Atom'}
        entries = root.findall('atom:entry', ns)
        
        import re
        items = []
        for entry in entries[:max_results]:
            title_node = entry.find("atom:title", ns)
            link_node = entry.find("atom:link", ns)
            pub_node = entry.find("atom:published", ns)
            
            title_text = title_node.text if title_node is not None else ""
            # 清理標題中的 HTML 標籤（Google Alerts 常會把關鍵字加上 <b> 標籤）
            title_text = re.sub(r'<[^>]+>', '', title_text)
            
            raw_link = link_node.attrib.get("href") if link_node is not None else ""
            # 解析 Google Alerts 跳轉網址以取得真實網址
            real_link = raw_link
            if "google.com/url" in raw_link:
                try:
                    parsed = urllib.parse.urlparse(raw_link)
                    qs = urllib.parse.parse_qs(parsed.query)
                    if 'url' in qs:
                        real_link = qs['url'][0]
                except Exception as e:
                    print(f"  [警訊] 解析快訊跳轉網址失敗 ({raw_link}): {e}")
            
            pub_date = pub_node.text if pub_node is not None else ""
            
            items.append({
                "title": title_text.strip(),
                "link": real_link.strip(),
                "pub_date": pub_date
            })
        return items
    except Exception as e:
        print(f"  [警訊] 抓取 Google 快訊 RSS 失敗: {e}")
        return []

def fetch_news_with_fallback(base_query, max_results=3):
    """
    先嘗試當天 (24小時)，若無結果則自動 fallback 至過去 3 天，以防止假日新聞較少的情況
    """
    query_1d = f"{base_query} when:1d"
    items = fetch_google_news_rss(query_1d, max_results)
    if not items:
        print(f"  [資訊] 過去 24 小時內無新聞，擴大搜尋範圍至過去 3 天...")
        query_3d = f"{base_query} when:3d"
        items = fetch_google_news_rss(query_3d, max_results)
    return items

def decode_url(google_url):
    """
    使用 googlenewsdecoder 庫解密 Google News 的跳轉 URL 以獲取原始發布者的真實連結
    """
    try:
        decoded_info = gnewsdecoder(google_url, interval=1)
        if decoded_info.get("status"):
            return decoded_info["decoded_url"]
    except Exception as e:
        print(f"  [警訊] URL 解碼失敗: {e}")
    return google_url

def extract_webpage_text(url, session=None):
    """
    抓取指定網頁，過濾無效的 CSS/JS，僅回傳網頁的純文字內容。
    如果是 DIGITIMES 的新聞且提供了登入的 session，將使用該 session 下載以獲取會員全文。
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7"
    }
    try:
        if "digitimes.com.tw" in url and session is not None:
            response = session.get(url, timeout=5)
        else:
            response = requests.get(url, headers=headers, timeout=5, verify=True)
            
        response.raise_for_status()
        
        if response.encoding == 'ISO-8859-1' or response.encoding is None:
            response.encoding = response.apparent_encoding
            
        soup = BeautifulSoup(response.text, 'html.parser')
        
        for element in soup(["script", "style", "nav", "footer", "header", "aside", "noscript"]):
            element.extract()
            
        text = soup.get_text(separator='\n')
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        cleaned_text = '\n'.join(lines)
        
        return cleaned_text[:20000]
    except Exception as e:
        print(f"  [警訊] 抓取網頁內文失敗 ({url}): {e}")
        return ""

def analyze_news_content(html_text):
    """
    使用 Gemini 2.5-flash 進行新聞的摘要、影響分析與結構化 JSON 輸出
    """
    prompt = f"""
請幫我分析以下這段網頁內容，它是一篇新聞報導。
網頁清理後的純文字如下：
---
{html_text}
---
請幫我提取並整理出：
1. 精簡主題標題 (title) - 根據新聞內容提煉出極簡短的主題標題(Topic)，讓讀者能馬上抓到重點，不要照抄冗長的原標題，字數限 5-15 字以內。
2. 去除雜訊後的新聞全文內文 (content_clean) - 提取真實新聞正文，排除網頁廣告、選單、版權宣告。
3. 新聞發布時間 (publish_time) - 如 YYYY-MM-DD HH:MM 格式，若無詳細時間則提取日期即可。
4. 內容摘要 (summary) - 內文敘述請精簡成 2-3 句話，直陳事實、數據與核心事件。
5. 華碩與筆電市場影響分析 (impact_analysis) - 評估該事件對「整個筆記型電腦 (PC/NB) 市場」與「華碩 (ASUS)」的關鍵影響，必須以「[AI觀點]」開頭，然後縮短為 1-2 句話。

請嚴格遵守回傳的 JSON 格式 Schema，所有文字均須以繁體中文 (Traditional Chinese) 回答。
"""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt,
                config={
                    'response_mime_type': 'application/json',
                    'response_schema': NewsItemAnalysis,
                }
            )
            return response.parsed
        except Exception as e:
            err_msg = str(e)
            if "503" in err_msg and attempt < max_retries - 1:
                print(f"  [警訊] Gemini 伺服器滿載 (503)，等待 5 秒後進行第 {attempt+2} 次重試...")
                time.sleep(5)
                continue
                
            if "429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg:
                print("  [警訊] Gemini API 額度已達上限 (429 RESOURCE_EXHAUSTED)，無法進行 AI 分析。")
            elif "400" in err_msg or "API_KEY_INVALID" in err_msg:
                print("  [警訊] Gemini API 金鑰無效，請檢查 .env 設定。")
            else:
                print(f"  [警訊] Gemini 分析失敗: {err_msg[:100]}...")
            return None

def analyze_news_from_title(title, category):
    """
    當無法抓取網頁全文時，使用新聞標題結合 Gemini 2.5-flash 的知識庫進行背景與影響摘要
    """
    prompt = f"""
您是一位專業的科技與財經分析師。
目前我們有一篇新聞的標題為：「{title}」，屬於「{category}」主題。
由於目前網路抓取受到限制，請您光憑這個「新聞標題」，並結合您的背景知識，為我們撰寫：
1. 精簡主題標題 (title) - 根據原新聞標題提煉出極簡短的主題標題(Topic)，讓讀者能馬上抓到重點，不要照抄冗長的原標題，字數限 5-15 字以內。
2. 新聞重點背景說明 (summary) - 內文敘述請精簡成 2-3 句話，直陳事實與重點。
3. 華碩與筆電市場影響分析 (impact_analysis) - 評估本則新聞對「筆電市場 (PC/NB)」與「華碩 (ASUS)」的潛在影響，必須以「[AI觀點]」開頭，然後縮短為 1-2 句話。

要求：
- 語氣需客觀、專業且肯定，不要使用「因為無法抓取內文」、「根據標題猜測」、「AI 預測」或「由於限制」等任何與系統限制相關的免責字句，直接產出分析內容即可。

請嚴格遵守回傳的 JSON 格式 Schema，所有文字均須以繁體中文 (Traditional Chinese) 回答。
"""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt,
                config={
                    'response_mime_type': 'application/json',
                    'response_schema': TitleOnlyAnalysis,
                }
            )
            return response.parsed
        except Exception as e:
            err_msg = str(e)
            if "503" in err_msg and attempt < max_retries - 1:
                print(f"  [警訊] Gemini 伺服器滿載 (503)，等待 5 秒後進行第 {attempt+2} 次重試...")
                time.sleep(5)
                continue
                
            if "429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg:
                print("  [警訊] Gemini API 額度已達上限 (429 RESOURCE_EXHAUSTED)，無法進行 AI 標題分析。")
            elif "400" in err_msg or "API_KEY_INVALID" in err_msg:
                print("  [警訊] Gemini API 金鑰無效，請檢查 .env 設定。")
            else:
                print(f"  [警訊] Gemini 標題分析失敗: {err_msg[:100]}...")
            return None

def generate_html_dashboard(excel_path, html_path):
    """
    從 Excel 讀取最新新聞，並生成具有美學動態效果的 HTML 互動看板。
    會過濾無效空白行，並動態將週數呈現在標題上。
    """
    try:
        if not os.path.exists(excel_path):
            return
            
        df = pd.read_excel(excel_path, sheet_name="News")
        if df.empty:
            return
            
        # 讀取最前 60 筆以確保在過濾空白行後能有足夠的資料量展示
        latest_df = df.head(60).fillna("")
        
        # 1. 決定動態週數標題：找到最新一筆有效的新聞列
        valid_rows = latest_df[latest_df["Topic"].astype(str).str.strip() != ""]
        if not valid_rows.empty:
            first_row = valid_rows.iloc[0]
            try:
                year_val = int(first_row.get("Year", datetime.now().year))
            except:
                year_val = datetime.now().year
            try:
                week_val = int(first_row.get("Week", datetime.now().isocalendar()[1]))
            except:
                week_val = datetime.now().isocalendar()[1]
        else:
            year_val = datetime.now().year
            week_val = datetime.now().isocalendar()[1]
            
        header_title = f"{year_val}wk{week_val:02d} News Summary"
        
        # 只保留與最新一筆週數相同的資料，避免混入上一週的新聞
        def is_current_week(r):
            try:
                return int(float(r.get("Year", 0))) == year_val and int(float(r.get("Week", 0))) == week_val
            except:
                return False
        latest_df = latest_df[latest_df.apply(is_current_week, axis=1)]
        
        categories = ["國際經濟", "PC產業", "上游廠商動態", "各品牌廠動態"]
        news_data = {cat: [] for cat in categories}
        
        # 2. 遍歷數據並進行有效性過濾
        for _, row in latest_df.iterrows():
            cat = str(row.get("Category", "")).strip()
            topic = str(row.get("Topic", "")).strip()
            
            # 過濾空白無效列 (若 Topic 為空則跳過)
            if not topic or topic.lower() == "nan":
                continue
                
            if cat in news_data:
                # 限制每個分類在前端只取最前 5 筆展示以求乾淨緊湊
                if len(news_data[cat]) >= 5:
                    continue
                    
                raw_time = str(row.get("時間", ""))
                clean_time = raw_time.split("\n")[0] if "\n" in raw_time else raw_time
                
                # === 使用者指定的標題客製化覆寫 ===
                title_text = str(topic)
                if "OECD" in title_text and cat == "國際經濟":
                    title_text = "OECD下修今年全球成長至2.8%，能源供應成關鍵變數"
                
                news_data[cat].append({
                    "title": html.escape(title_text),
                    "summary": html.escape(str(row.get("Content", ""))).replace("\n", "<br>"),
                    "impact": html.escape(str(row.get("對於筆電市場/華碩的影響", ""))).replace("\n", "<br>"),
                    "link": str(row.get("Link", "")),
                    "time": html.escape(clean_time),
                    "full_content": html.escape(str(row.get("完整內容", ""))).replace("\n", "<br>")
                })

        # 生成各類別的新聞條目 HTML
        cat_icons = {
            "國際經濟": "📈",
            "PC產業": "💻",
            "上游廠商動態": "⚙️",
            "各品牌廠動態": "🏷️"
        }
        
        # 使用 Gemini 為每個分類生成一句話標題
        def generate_section_headline(cat_name, items):
            """呼叫 Gemini 根據該分類下所有新聞標題，生成一句話總結標題"""
            if not items:
                return cat_name
            titles = "、".join([item["title"] for item in items])
            prompt = f"""你是一位專業的科技產業分析師。
以下是「{cat_name}」分類下本週的所有新聞標題：
{titles}

請用一句話（20-35字以內）歸納這些新聞的核心主題，作為本週此分類的總標題。
要求：直接輸出標題文字，不要加引號、不要加標點符號結尾、不要解釋。"""
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = client.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=prompt
                    )
                    headline = response.text.strip().strip('"').strip('「').strip('」').strip('。').strip()
                    return headline if headline else cat_name
                except Exception as e:
                    if "503" in str(e) and attempt < max_retries - 1:
                        print(f"  [警訊] 生成 {cat_name} 標題時伺服器擁擠 (503)，等待 5 秒後進行第 {attempt+2} 次重試...")
                        time.sleep(5)
                        continue
                    print(f"  [警訊] 生成 {cat_name} 標題失敗: {e}")
                    return cat_name
        
        # 使用 Gemini 合併所有 AI 觀點為一段總結
        def generate_consolidated_impact(cat_name, items):
            """呼叫 Gemini 將該分類下所有新聞的 AI 觀點合併為一段總結"""
            if not items:
                return ""
            impacts = [item["impact"] for item in items if item["impact"] and item["impact"].strip() and item["impact"].strip().lower() != "nan"]
            if not impacts:
                return ""
            all_impacts = "\n".join(impacts)
            prompt = f"""你是一位專業的筆電產業分析師。
以下是「{cat_name}」分類下本週所有新聞的個別 AI 觀點（對筆電市場/華碩的影響）：
---
{all_impacts}
---

請將上述所有觀點整合為一段極度精簡的總結分析（1-3句話以內，用字精煉），說明這些新聞綜合來看對「筆記型電腦市場」與「華碩(ASUS)」的關鍵影響。
要求：以「[AI觀點]」開頭，語氣客觀專業且科技感，使用繁體中文。"""
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = client.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=prompt
                    )
                    return response.text.strip() if response.text else ""
                except Exception as e:
                    if "503" in str(e) and attempt < max_retries - 1:
                        print(f"  [警訊] 合併 {cat_name} AI觀點時伺服器擁擠 (503)，等待 5 秒後進行第 {attempt+2} 次重試...")
                        time.sleep(5)
                        continue
                    print(f"  [警訊] 合併 {cat_name} AI觀點失敗: {e}")
                    # fallback: 如果失敗，只取前兩個觀點，移除重複的標籤並以條列呈現
                    cleaned = []
                    for imp in impacts[:2]:
                        imp = imp.replace("[AI觀點]", "").strip()
                        cleaned.append(f"<li style='margin-bottom: 0.3rem;'>{imp}</li>")
                    return "[AI觀點] <ul style='padding-left: 1.2rem; margin-top: 0.4rem; margin-bottom: 0;'>" + "".join(cleaned) + "</ul>"
        
        # 為每個分類生成標題與總結 AI 觀點
        print("-> 正在為各分類生成標題與 AI 觀點總結...")
        section_headlines = {}
        section_impacts = {}
        for cat in categories:
            items = news_data[cat]
            if items:
                section_headlines[cat] = generate_section_headline(cat, items)
                print(f"  [{cat}] 標題: {section_headlines[cat]}")
                section_impacts[cat] = generate_consolidated_impact(cat, items)
                time.sleep(1)
            else:
                section_headlines[cat] = cat
                section_impacts[cat] = ""
        
        # 4欄版面：將所有分類並排（各分類獨立配色）
        def build_four_columns_html():
            cat_colors = {
                "國際經濟":    {"main": "#b45309", "bg": "rgba(180,83,9,0.06)",   "border": "rgba(180,83,9,0.2)"},
                "PC產業":     {"main": "#0369a1", "bg": "rgba(3,105,161,0.06)",   "border": "rgba(3,105,161,0.2)"},
                "上游廠商動態": {"main": "#6d28d9", "bg": "rgba(109,40,217,0.06)", "border": "rgba(109,40,217,0.2)"},
                "各品牌廠動態": {"main": "#047857", "bg": "rgba(4,120,87,0.06)",   "border": "rgba(4,120,87,0.2)"},
            }
            col_html = '<div class="news-grid">'
            for cat in categories:
                items = news_data[cat]
                icon = cat_icons.get(cat, '📰')
                headline = section_headlines.get(cat, cat)
                impact_summary = section_impacts.get(cat, "")
                c = cat_colors.get(cat, cat_colors["PC產業"])
                main_color = c["main"]
                bg_color = c["bg"]
                border_color = c["border"]

                col_html += f'<div class="news-col" style="--cat-color:{main_color}; --cat-bg:{bg_color}; --cat-border:{border_color};">'

                # ── 欄標題區
                col_html += f'''
                <div class="cat-header">
                    <div class="cat-meta">
                        <span class="cat-icon">{icon}</span>
                        <span class="cat-label">{cat}</span>
                        <span class="cat-count">{len(items)}</span>
                    </div>
                    <div class="cat-summary">{headline}</div>
                </div>
                '''

                # ── AI 觀點區塊
                if impact_summary:
                    impact_clean = impact_summary.replace("[AI觀點]", "").replace("**", "").strip()
                    col_html += f'''
                    <div class="ai-insight">
                        <div class="ai-label">💡 AI 觀點</div>
                        <div class="ai-text">{impact_clean}</div>
                    </div>
                    '''

                # ── 新聞列表
                if not items:
                    col_html += '<p class="no-news">此分類本週尚無新聞。</p>'
                else:
                    for item in items:
                        link_btn = ""
                        if item["link"] and item["link"].startswith("http"):
                            link_btn = f'<a href="{item["link"]}" target="_blank" class="link-btn">🔗 原文</a>'
                        time_chip = ""
                        t = item.get("time", "")
                        if t and t != "nan" and t.strip():
                            time_chip = f'<span class="news-time">{t[:10]}</span>'
                        col_html += f'''
                        <div class="news-item">
                            <div class="news-row">
                                <h4 class="news-title">{item['title']}</h4>
                                {link_btn}
                            </div>
                            <p class="news-body">{item['summary']}</p>
                            {time_chip}
                        </div>
                        '''

                col_html += '</div>'
            col_html += '</div>'
            return col_html
            
        grid_html = build_four_columns_html()
        history_html = ""
        try:
            import glob
            history_files = glob.glob(os.path.join(os.path.dirname(html_path), "*wk*.html"))
            weeks = []
            for f in history_files:
                basename = os.path.basename(f)
                if basename.endswith(".html") and "wk" in basename:
                    w = basename.replace(".html", "")
                    if w not in weeks:
                        weeks.append(w)
            
            # 確保當前週數也在清單中
            current_wk_str = f"{year_val}wk{week_val:02d}"
            if current_wk_str not in weeks:
                weeks.append(current_wk_str)
                
            # 排序（新的在最前面）
            weeks.sort(reverse=True)
            
            if weeks:
                pills = []
                for w in weeks[:8]:  # 最多顯示 8 週
                    active_cls = "week-pill active" if w == current_wk_str else "week-pill"
                    val = "index.html" if w == current_wk_str else f"{w}.html"
                    pills.append(f'<a href="{val}" class="{active_cls}">{w}</a>')
                history_html = f'<div class="history-nav">{"".join(pills)}</div>'
        except Exception as e:
            print(f"產生歷史選單錯誤: {e}")

        # 完整的 HTML 看板模板 (NotebookLM Infographic 風格)
        html_template = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{header_title}</title>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;500;700&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}

        body {{
            background: #f0f4f8;
            color: #1e293b;
            font-family: 'Noto Sans TC', system-ui, sans-serif;
            line-height: 1.5;
            min-height: 100vh;
        }}

        /* ── Header ── */
        .page-header {{
            background: #ffffff;
            border-bottom: 1px solid #e2e8f0;
            padding: 0.75rem 1.5rem;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 1rem;
            position: sticky;
            top: 0;
            z-index: 100;
            box-shadow: 0 1px 4px rgba(0,0,0,0.07);
        }}
        .header-brand {{
            display: flex;
            align-items: center;
            gap: 0.6rem;
            flex-shrink: 0;
        }}
        .header-logo {{
            width: 28px; height: 28px;
            background: #1e40af;
            border-radius: 6px;
            display: flex; align-items: center; justify-content: center;
            color: #fff; font-size: 0.85rem; font-weight: 800;
        }}
        .header-title {{
            font-size: 1rem;
            font-weight: 700;
            color: #1e293b;
        }}
        .header-subtitle {{
            font-size: 0.75rem;
            color: #64748b;
        }}
        .history-nav {{
            display: flex;
            align-items: center;
            gap: 0.35rem;
            flex-wrap: wrap;
        }}
        .week-pill {{
            display: inline-block;
            padding: 0.25rem 0.65rem;
            border-radius: 20px;
            font-size: 0.72rem;
            font-weight: 600;
            text-decoration: none;
            border: 1px solid #cbd5e1;
            color: #475569;
            background: #f8fafc;
            transition: all 0.15s;
            white-space: nowrap;
        }}
        .week-pill:hover {{ border-color: #1e40af; color: #1e40af; background: #eff6ff; }}
        .week-pill.active {{ background: #1e40af; color: #fff; border-color: #1e40af; }}
        .header-date {{ font-size: 0.72rem; color: #94a3b8; flex-shrink: 0; }}

        /* ── Grid ── */
        .news-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 1rem;
            padding: 1rem 1.25rem 1.5rem;
            max-width: 1800px;
            margin: 0 auto;
        }}
        .news-col {{
            display: flex;
            flex-direction: column;
            gap: 0.75rem;
            min-width: 0;
        }}

        /* ── Cat header ── */
        .cat-header {{
            background: #fff;
            border-radius: 10px;
            padding: 0.8rem 1rem;
            border: 1px solid #e2e8f0;
            border-top: 3px solid var(--cat-color);
        }}
        .cat-meta {{ display: flex; align-items: center; gap: 0.45rem; margin-bottom: 0.4rem; }}
        .cat-icon {{ font-size: 1rem; line-height: 1; }}
        .cat-label {{ font-size: 0.7rem; font-weight: 800; letter-spacing: 1px; color: var(--cat-color); }}
        .cat-count {{
            margin-left: auto;
            background: var(--cat-color); color: #fff;
            font-size: 0.65rem; font-weight: 700;
            padding: 0.1rem 0.45rem; border-radius: 12px;
        }}
        .cat-summary {{ font-size: 0.82rem; font-weight: 600; color: #334155; line-height: 1.45; }}

        /* ── AI insight ── */
        .ai-insight {{
            background: var(--cat-bg);
            border: 1px solid var(--cat-border);
            border-left: 3px solid var(--cat-color);
            border-radius: 8px;
            padding: 0.7rem 0.85rem;
        }}
        .ai-label {{ font-size: 0.65rem; font-weight: 800; color: var(--cat-color); letter-spacing: 0.8px; margin-bottom: 0.3rem; }}
        .ai-text {{ font-size: 0.8rem; color: #374151; line-height: 1.6; }}

        /* ── News items ── */
        .news-item {{
            background: #fff;
            border: 1px solid #e2e8f0;
            border-left: 3px solid var(--cat-color);
            border-radius: 8px;
            padding: 0.7rem 0.85rem;
            transition: box-shadow 0.15s, transform 0.15s;
        }}
        .news-item:hover {{ box-shadow: 0 4px 12px rgba(0,0,0,0.09); transform: translateY(-1px); }}
        .news-row {{
            display: flex; align-items: flex-start;
            justify-content: space-between; gap: 0.4rem; margin-bottom: 0.35rem;
        }}
        .news-title {{ font-size: 0.88rem; font-weight: 700; color: #1e293b; line-height: 1.4; flex: 1; }}
        .link-btn {{
            flex-shrink: 0;
            background: var(--cat-bg); color: var(--cat-color);
            border: 1px solid var(--cat-border);
            padding: 0.15rem 0.45rem; border-radius: 4px;
            font-size: 0.63rem; font-weight: 600;
            text-decoration: none; white-space: nowrap; transition: all 0.15s;
        }}
        .link-btn:hover {{ background: var(--cat-color); color: #fff; }}
        .news-body {{
            font-size: 0.78rem; color: #4b5563; line-height: 1.65;
            display: -webkit-box; -webkit-line-clamp: 5;
            -webkit-box-orient: vertical; overflow: hidden;
        }}
        .news-time {{
            display: inline-block; margin-top: 0.4rem;
            font-size: 0.62rem; color: #94a3b8;
            background: #f1f5f9; padding: 0.1rem 0.4rem; border-radius: 3px;
        }}
        .no-news {{ font-size: 0.8rem; color: #94a3b8; padding: 0.5rem; }}

        /* ── Responsive ── */
        @media (max-width: 1100px) {{ .news-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
        @media (max-width: 640px) {{
            .news-grid {{ grid-template-columns: 1fr; padding: 0.75rem; }}
            .page-header {{ flex-wrap: wrap; }}
        }}
        @media print {{ body {{ background: #fff; }} .page-header {{ box-shadow: none; position: static; }} }}
    </style>
</head>
<body>
    <div class="page-header">
        <div class="header-brand">
            <div class="header-logo">N</div>
            <div>
                <div class="header-title">ASUS News Intelligence</div>
                <div class="header-subtitle">{year_val}wk{week_val:02d} · PC 產業週報</div>
            </div>
        </div>
        {history_html}
        <div class="header-date">{datetime.now().strftime('%Y-%m-%d')} 更新</div>
    </div>
    {grid_html}
</body>
</html>
"""
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_template)
            
        current_wk_str = f"{year_val}wk{week_val:02d}"
        archive_path = os.path.join(os.path.dirname(html_path), f"{current_wk_str}.html")
        with open(archive_path, "w", encoding="utf-8") as f:
            f.write(html_template)
            
        print(f"成功更新 HTML 看板網頁：{html_path}")
        print(f"成功儲存歷史歸檔：{archive_path}")
        
        # 4. 同步更新所有歷史歸檔檔案的頂部導覽列，確保連結能往返切換
        import re
        if 'history_files' in locals() and history_files:
            print("-> 正在同步更新所有歷史歸檔網頁的導覽列...")
            for f_path in history_files:
                # 排除剛寫入的 archive_path，避免重複讀寫
                if os.path.abspath(f_path) == os.path.abspath(archive_path):
                    continue
                try:
                    with open(f_path, "r", encoding="utf-8") as f_old:
                        content_old = f_old.read()
                    
                    # 用最新產生的 history_html 替換舊有的 history-nav 導覽列
                    new_content_old = re.sub(
                        r'<div class="history-nav">.*?</div>',
                        history_html,
                        content_old,
                        flags=re.DOTALL
                    )
                    
                    # 順便更新每個舊頁面的 week-pill active 狀態，確保使用者清楚當前停留在哪一頁
                    basename_no_ext = os.path.basename(f_path).replace(".html", "")
                    
                    # 先將所有 pill 降為普通樣式，再把對應此舊週次檔案的按鈕設為 active
                    new_content_old = new_content_old.replace('week-pill active', 'week-pill')
                    new_content_old = new_content_old.replace(
                        f'href="{basename_no_ext}.html" class="week-pill"',
                        f'href="{basename_no_ext}.html" class="week-pill active"'
                    )
                    
                    with open(f_path, "w", encoding="utf-8") as f_old:
                        f_old.write(new_content_old)
                except Exception as ex:
                    print(f"  [警訊] 同步更新歷史檔案 {os.path.basename(f_path)} 導覽列失敗: {ex}")
    except Exception as e:
        print(f"  [警訊] 生成 HTML 看板時出錯: {e}")

def auto_deploy_to_github():
    """
    使用 subprocess 自動將最新的 HTML 網頁推送到 GitHub Pages。
    自動尋找包含 .git 的工作區根目錄。
    """
    import subprocess
    # 尋找包含 .git 的目錄（優先檢查 NEWS 目錄，若無則檢查上層目錄）
    base_dir = r"D:\ASUS\Anti-NotebookLM\NEWS"
    repo_dir = base_dir
    if not os.path.exists(os.path.join(base_dir, ".git")):
        parent_dir = os.path.dirname(base_dir)
        if os.path.exists(os.path.join(parent_dir, ".git")):
            repo_dir = parent_dir
            
    print("\n" + "=" * 60)
    print(" 🚀 啟動一鍵自動發布至 GitHub Pages...")
    print(f"  [工作目錄] {repo_dir}")
    print("=" * 60)
    
    try:
        # 1. git add .
        res_add = subprocess.run(["git", "add", "."], cwd=repo_dir, capture_output=True, encoding='utf-8', errors='replace')
        if res_add.returncode != 0:
            err_msg = (res_add.stderr or "").strip()
            print(f"  [警訊] Git add 失敗: {err_msg}")
            return
            
        # 2. git commit -m
        commit_msg = f"Auto update news dashboard ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
        res_commit = subprocess.run(["git", "commit", "-m", commit_msg], cwd=repo_dir, capture_output=True, encoding='utf-8', errors='replace')
        
        stdout_str = res_commit.stdout or ""
        stderr_str = res_commit.stderr or ""
        
        if "nothing to commit" in stdout_str or "nothing to commit" in stderr_str:
            print("  [資訊] 本次變更內容無差異或已經是最新狀態。")
        elif res_commit.returncode != 0:
            print(f"  [資訊] Git commit 狀態: {stdout_str.strip() or stderr_str.strip()}")
            
        # 3. git push
        print("  -> 正在推送到雲端 GitHub...")
        res_push = subprocess.run(["git", "push"], cwd=repo_dir, capture_output=True, encoding='utf-8', errors='replace')
        if res_push.returncode != 0:
            # 嘗試設定 upstream 並推送到 origin main
            res_push = subprocess.run(["git", "push", "-u", "origin", "main"], cwd=repo_dir, capture_output=True, encoding='utf-8', errors='replace')
            if res_push.returncode != 0 and ("rejected" in res_push.stderr or "fetch first" in res_push.stderr):
                print("  [資訊] 偵測到雲端倉庫有舊歷史紀錄，正在進行首次強制對齊發布...")
                res_push = subprocess.run(["git", "push", "-u", "origin", "main", "--force"], cwd=repo_dir, capture_output=True, encoding='utf-8', errors='replace')
            
        if res_push.returncode == 0:
            print("  [成功] 網頁已成功自動推送到 GitHub！線上網址將在幾秒內更換為最新內容。")
        else:
            err_msg = (res_push.stderr or "").strip()
            print(f"  [警訊] Git push 失敗。請確認本機是否已連結 GitHub 倉庫。")
            print(f"  [錯誤細節] {err_msg}")
    except Exception as e:
        print(f"  [警訊] 自動發布過程發生異常: {e}")


def append_to_excel(file_path, new_rows):
    """
    使用 openpyxl 將整理完的資料安全地寫入 Excel 的最上方，保留其它 Sheet 的內容與格式。
    會自動檢查並動態插入「對於筆電市場/華碩的影響」欄位，維持與現有格式的相容。
    在寫入被 Excel 鎖定時提供互動式提示。
    """
    # 完美對接的全新 headers (將新欄位設定在第九欄 I 欄，對應原本的 Unnamed: 8)
    headers = ['Year', 'Week', 'Category', 'Topic', 'Content', 'Link', '時間', '完整內容', '對於筆電市場/華碩的影響', 'Unnamed: 9']
    
    # 檢查 Excel 目錄是否存在，若不存在則建立
    dir_path = os.path.dirname(file_path)
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)

    if not os.path.exists(file_path):
        print(f"Excel 檔案不存在，正在全新建立：{file_path}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "News"
        ws.append(headers)
    else:
        wb = openpyxl.load_workbook(file_path)
        if "News" not in wb.sheetnames:
            ws = wb.create_sheet("News")
            ws.append(headers)
        else:
            ws = wb["News"]
            
            # 動態檢測：檢查第 9 欄 (I 欄) 是否為「對於筆電市場/華碩的影響」
            # 如果不是 (原本是空的或 Unnamed: 8)，直接設定第 9 欄第一行為該 Header
            if ws.cell(row=1, column=9).value != '對於筆電市場/華碩的影響':
                print("  [資訊] 偵測到 Excel 第 9 欄 (I 欄) 尚未命名，將其命名為「對於筆電市場/華碩的影響」...")
                ws.cell(row=1, column=9, value='對於筆電市場/華碩的影響')
            
    # 將新資料插入在工作表的最前面（Header 正下方的第二行）
    num_new_rows = len(new_rows)
    if num_new_rows > 0:
        ws.insert_rows(2, num_new_rows)
        
        # 從第二行開始，依次填入新抓取的新聞
        for idx, row_data in enumerate(new_rows):
            current_row = 2 + idx
            for col_idx, h in enumerate(headers, 1):
                if h.startswith("Unnamed:"):
                    val = ""
                else:
                    val = row_data.get(h, "")
                ws.cell(row=current_row, column=col_idx, value=val)
        
    # 儲存工作簿並加入 PermissionError (檔案被 Excel 開啟中) 處理
    try:
        wb.save(file_path)
        print(f"\n[成功] 成功將 {len(new_rows)} 筆新聞資料寫入 Excel 檔案！")
    except PermissionError:
        print("\n" + "!"*60)
        print("[警告] 無法儲存 Excel 檔案！該檔案可能已被微軟 Excel 軟體開啟。")
        print("!"*60 + "\n")
        while True:
            ans = input("請先關閉該 Excel 檔案，然後輸入 'Y' 以重新嘗試存檔 (或輸入 'N' 放棄儲存): ")
            if ans.upper() == 'Y':
                try:
                    wb.save(file_path)
                    print("[成功] 存檔成功！")
                    break
                except PermissionError:
                    print("檔案仍被 Excel 鎖定，請關閉後再試。")
            elif ans.upper() == 'N':
                print("[取消] 放棄儲存本次抓取的新聞資料。")
                break

def main():
    print("=" * 60)
    print(" 每日新聞閱讀與 Excel 整理工具啟動")
    print(f" 執行時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f" Excel 目標：{EXCEL_PATH}")
    print("=" * 60)
    
    # 嘗試建立 DIGITIMES 會員 Session
    dgt_session = None
    if DIGITIMES_USER and DIGITIMES_PASSWORD:
        dgt_session = get_digitimes_session(DIGITIMES_USER, DIGITIMES_PASSWORD)
        
    # 計算當前的年份
    now = datetime.now()
    current_year = now.year

    # 嘗試讀取現有 Excel 中的歷史新聞，建立去重清單，並從最新週次 +1 決定本次週數
    existing_titles = set()
    existing_links = set()
    current_week = now.isocalendar()[1]  # 預設值：系統 ISO 週數
    if os.path.exists(EXCEL_PATH):
        try:
            df_history = pd.read_excel(EXCEL_PATH)
            if 'Topic' in df_history.columns:
                existing_titles = set(df_history['Topic'].dropna().astype(str).str.strip())
            if 'Link' in df_history.columns:
                existing_links = set(df_history['Link'].dropna().astype(str).str.strip())
            print(f"  [資訊] 成功載入歷史紀錄：包含 {len(existing_titles)} 篇已抓取新聞，將自動過濾重複。")
            # 從 Excel 最新年份的最大 Week +1 推算本次週數，並讓使用者確認
            if 'Week' in df_history.columns and 'Year' in df_history.columns:
                max_year_in_excel = int(df_history['Year'].dropna().astype(float).max())
                df_latest_yr = df_history[df_history['Year'].dropna().astype(float).astype(int) == max_year_in_excel]
                max_week_vals = df_latest_yr['Week'].dropna()
                if not max_week_vals.empty:
                    suggested_week = int(max_week_vals.astype(float).max()) + 1
                    suggested_year = max_year_in_excel
                    if suggested_week > 53:
                        suggested_week = 1
                        suggested_year += 1
                    print(f"\n  [確認] 偵測到 Excel 最新資料為 {max_year_in_excel}wk{int(max_week_vals.astype(float).max()):02d}")
                    print(f"  [確認] 建議本次寫入週次：{suggested_year}wk{suggested_week:02d}")
                    ans = input(f"  請確認週次（直接 Enter 使用建議值，或輸入數字如 27 覆蓋）：").strip()
                    if ans.isdigit():
                        current_week = int(ans)
                        current_year = suggested_year
                    else:
                        current_week = suggested_week
                        current_year = suggested_year
                    print(f"  [資訊] 本次寫入週次：{current_year}wk{current_week:02d}")
        except Exception as e:
            print(f"  [警訊] 無法讀取 Excel 歷史紀錄進行去重: {e}")
            
    collected_data = []
    
    for category_name, query_base in CATEGORIES.items():
        print(f"\n-> 正在搜尋主題：【{category_name}】")
        
        all_candidates = []
        if category_name == "國際經濟" and GOOGLE_ALERT_RSS_ECON:
            print(f"  [資訊] 偵測到 GOOGLE_ALERT_RSS_ECON，改從 Google 快訊 RSS 抓取新聞...")
            # 抓取最多 15 篇較多候選，以便挑選包含優先關鍵字的新聞
            all_candidates = fetch_google_alert_rss(GOOGLE_ALERT_RSS_ECON, max_results=15)
            
            # 優先權排序邏輯：優先處理「中國」、「美國」與宏觀的「全球經濟」，並降低「台灣本土企業/台股」新聞的優先度
            top_priority = ["中國", "美國", "美中", "聯準會", "Fed", "降息", "通膨"]
            macro_priority = ["全球經濟", "全球市場", "國際經濟", "全球成長"]
            exclude_keywords = ["董座", "董事長", "台股", "台廠", "三陽", "光陽", "台積電", "聯電", "鴻海"]
            
            def get_priority(item):
                title = item.get("title", "")
                # 1. 包含台灣本土企業/台股相關詞彙，優先度降到最低 (回傳 3)
                if any(x in title for x in exclude_keywords):
                    return 3
                # 2. 包含美中等關鍵字，頂級優先 (回傳 0)
                if any(k in title for k in top_priority):
                    return 0
                # 3. 包含宏觀全球經濟關鍵字，次級優先 (回傳 1)
                if any(k in title for k in macro_priority):
                    return 1
                # 4. 包含普通「全球」字眼，三級優先 (回傳 2)
                if "全球" in title:
                    return 2
                # 5. 其餘新聞 (回傳 3)
                return 3
            
            all_candidates.sort(key=get_priority)
            print(f"  [資訊] 已根據優先關鍵字 {top_priority} 與排除詞彙重新排序候選新聞順序...")
        else:
            # 先找 DIGITIMES (優先，最多取2篇以保留空間給其他來源)
            digitimes_query = f"{query_base} site:digitimes.com.tw"
            digitimes_items = fetch_news_with_fallback(digitimes_query, max_results=2)
            
            # 搜尋其他管道補充
            print(f"  [資訊] 搜尋其他管道新聞...")
            other_query = f"{query_base} -site:digitimes.com.tw"
            other_items = fetch_news_with_fallback(other_query, max_results=5)
            
            # 組合候選清單：DIGITIMES 排在前面優先處理
            all_candidates = digitimes_items + other_items
            
        if not all_candidates:
            print(f"  -> 過去 3 天內未搜尋到相關新聞。")
            continue
            
        print(f"  -> 找到 {len(all_candidates)} 篇候選新聞，開始進行抓取與篩選...")
        
        valid_count = 0
        
        for idx, item in enumerate(all_candidates, 1):
            if valid_count >= MAX_ARTICLES_PER_CATEGORY:
                break
                
            title = item["title"]
            google_url = item["link"]
            pub_date = item["pub_date"]
            
            print(f"  [{idx}/{len(all_candidates)}] 處理新聞：{title[:40]}...")
            
            if title.strip() in existing_titles:
                print(f"  [跳過] 這篇新聞之前已經抓取過了 (標題重複)")
                continue
            
            # 1. 解碼 Google News 轉址
            real_url = decode_url(google_url)

            if real_url.strip() in existing_links:
                print(f"  [跳過] 這篇新聞之前已經抓取過了 (網址重複)")
                continue

            # 過濾購物/零售/產品頁面（非新聞）
            _BLOCKED_DOMAINS = [
                'tw.buy.yahoo.com', 'buy.yahoo.com', 'shopping.pchome.com.tw',
                'momoshop.com.tw', 'shopee.tw', 'momo.dm', 'ecshop',
            ]
            _BLOCKED_URL_PATTERNS = ['/product/', '/products/', '/item/', '/goods/', 'goods.ruten']
            if any(d in real_url for d in _BLOCKED_DOMAINS) or any(p in real_url for p in _BLOCKED_URL_PATTERNS):
                print(f"  [跳過] 疑似購物/產品頁面，略過 ({real_url[:70]}...)")
                continue

            # 2. 爬取內文純文字 (如果為 DIGITIMES 新聞會自動使用登入後的 session)
            web_text = extract_webpage_text(real_url, session=dgt_session)

            if not web_text or len(web_text.strip()) < 100:
                print(f"  [跳過] 無實質新聞內文可供摘要 ({title[:30]}...)")
                continue

            # 3. 呼叫 Gemini 進行整理與摘要
            analysis = analyze_news_content(web_text)
            time.sleep(1)

            if analysis:
                # 優先使用 AI 提煉的精簡標題，若空則 fallback 到原始標題
                raw_fallback = title.split(" - ")[0] if " - " in title else title
                topic = analysis.title.strip() if analysis.title and analysis.title.strip() else raw_fallback
                content = analysis.summary
                impact = analysis.impact_analysis
                pub_time = analysis.publish_time
                full_content = analysis.content_clean
            else:
                print(f"  [跳過] AI 分析失敗，略過此篇新聞 ({title[:30]}...)")
                continue

            # 過濾 UUID 式標題 (含連字符十六進位碼，如 0C813AA6-A7DD-41EA-975C)
            import re as _re
            if _re.search(r'[0-9A-Fa-f]{6,}-[0-9A-Fa-f]{4,}-[0-9A-Fa-f]{4,}', topic):
                print(f"  [跳過] 標題為系統編碼，非正常新聞 ({topic[:50]})")
                continue
            # 過濾產品規格式標題 (含記憶體/儲存規格，如 8G+16G/2TB SSD/Win11)
            if _re.search(r'(?:/\d+G[B]?|\d+GB|SSD|Win11|Win10|PCIe\s|DDR\d)', topic, _re.IGNORECASE):
                print(f"  [跳過] 標題含產品規格，疑似產品頁面 ({topic[:50]})")
                continue
                
            # 整理為 Excel 列結構
            row_data = {
                "Year": current_year,
                "Week": current_week,
                "Category": category_name,
                "Topic": topic,
                "Content": content,
                "對於筆電市場/華碩的影響": impact,
                "Link": real_url,
                "時間": pub_time,
                "完整內容": full_content
            }
            collected_data.append(row_data)
            
            # 即時更新去重清單，避免同一次執行中不同分類抓到一樣的新聞
            existing_titles.add(title.strip())
            existing_links.add(real_url.strip())
            
            valid_count += 1
            
    if collected_data:
        # 將資料追加寫入 Excel
        append_to_excel(EXCEL_PATH, collected_data)
        # 產生 HTML (輸出為 index.html 與 week_num.html)
        print("\n-> 正在產生 HTML 報表...")
        local_index_path = r"D:\ASUS\Anti-NotebookLM\NEWS\index.html"
        generate_html_dashboard(EXCEL_PATH, local_index_path)
        
        try:
            webbrowser.open(local_index_path)
            print("\n[成功] 已在瀏覽器中自動為您開啟「ASUS 新聞情報看板」網頁！")
        except Exception as e:
            print(f"  [警訊] 自動開啟網頁時出錯: {e}")
            
        # 自動推送到 GitHub
        auto_deploy_to_github()
    else:
        print("\n[結束] 今日沒有收集到任何新聞資料。")
        
    print("\n" + "=" * 60)
    print(" 程式執行完畢")
    print("=" * 60)

if __name__ == "__main__":
    main()
